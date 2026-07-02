"""
Lv10 - Excelレポート出力（openpyxl）
=====================================
CSVではなく「そのまま上司に提出できるExcelレポート」を Python で生成する。

CSVとの違い:
  - 複数シートを持てる（生データ＋集計を1ファイルに）
  - 書式（色・罫線・列幅・数値フォーマット）を設定できる
  - オートフィルタ・ウィンドウ枠の固定など「見る人に優しい」機能を付けられる

実務では「CSVで渡すと結局誰かがExcelで整形している」ことが多い。
その整形ごと自動化してしまうのがこのレベルのテーマ。

実行方法:
    pip install -r requirements.txt
    python main.py
"""

from datetime import datetime
from pathlib import Path

# openpyxl: Excel (.xlsx) ファイルを読み書きする定番ライブラリ
# pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 0. サンプルデータ
# ============================================================
# Lv09 のスクレイピング結果を想定したデータ。
# 実際には Lv09 の scraper が返す list[dict] をそのまま渡せる。

BOOKS = [
    {"title": "A Light in the Attic", "price": 51.77, "rating": 3, "stock": 22},
    {"title": "Tipping the Velvet", "price": 53.74, "rating": 1, "stock": 20},
    {"title": "Soumission", "price": 50.10, "rating": 1, "stock": 20},
    {"title": "Sharp Objects", "price": 47.82, "rating": 4, "stock": 20},
    {"title": "Sapiens: A Brief History", "price": 54.23, "rating": 5, "stock": 20},
    {"title": "The Requiem Red", "price": 22.65, "rating": 1, "stock": 19},
    {"title": "The Dirty Little Secrets", "price": 33.34, "rating": 4, "stock": 19},
    {"title": "The Coming Woman", "price": 17.93, "rating": 3, "stock": 19},
    {"title": "The Boys in the Boat", "price": 22.60, "rating": 4, "stock": 19},
    {"title": "The Black Maria", "price": 52.15, "rating": 1, "stock": 19},
    {"title": "Starving Hearts", "price": 13.99, "rating": 2, "stock": 19},
    {"title": "Shakespeare's Sonnets", "price": 20.66, "rating": 4, "stock": 19},
]

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ============================================================
# 1. スタイル定義 ─ 使い回す書式は先に作っておく
# ============================================================
# openpyxl では「セルに1つずつ書式を設定する」ため、
# よく使う書式をモジュール定数として定義しておくとコードが読みやすい。

# ヘッダー行: 濃い青背景 + 白太字
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# 罫線: 全セル共通の細線
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 中央揃え
CENTER = Alignment(horizontal="center", vertical="center")


# ============================================================
# 2. 生データシートを作る
# ============================================================

def write_data_sheet(wb: Workbook, books: list[dict]) -> None:
    """
    書籍一覧をシートに書き込む。

    ポイント:
      - ws.cell(row=行, column=列, value=値) でセルに書き込む（行・列は1始まり）
      - ws.append(リスト) なら1行分をまとめて追加できる
    """
    # Workbook 作成直後に自動で入っているシートを流用し、名前を付ける
    ws = wb.active
    ws.title = "書籍一覧"

    # --- ヘッダー行 ---
    headers = ["No.", "タイトル", "価格(£)", "星評価", "在庫数"]
    ws.append(headers)

    # ヘッダーに書式を適用（1行目の各セルをループ）
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # --- データ行 ---
    # enumerate(..., start=2) で「書き込み先の行番号」を同時に得る
    for row_num, book in enumerate(books, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1)          # 連番
        ws.cell(row=row_num, column=2, value=book["title"])
        ws.cell(row=row_num, column=3, value=book["price"])
        ws.cell(row=row_num, column=4, value=book["rating"])
        ws.cell(row=row_num, column=5, value=book["stock"])

        # 行全体に罫線
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = THIN_BORDER

        # 価格列は小数2桁表示にする（数値フォーマット）
        ws.cell(row=row_num, column=3).number_format = "0.00"

        # 星1（低評価）の行はタイトルを赤字にして目立たせる ─ 条件による書式分け
        if book["rating"] == 1:
            ws.cell(row=row_num, column=2).font = Font(color="C00000")

    # --- 列幅の調整 ---
    # 列幅は「文字数ベース」の単位。タイトル列だけ広くする
    widths = [6, 40, 10, 8, 8]
    for col_num, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # --- 見る人に優しい仕上げ ---
    ws.auto_filter.ref = ws.dimensions   # ヘッダーにオートフィルタを付ける
    ws.freeze_panes = "A2"               # 1行目を固定（スクロールしてもヘッダーが残る）

    print(f"  [書籍一覧] {len(books)} 件を書き込みました")


# ============================================================
# 3. 集計シートを作る
# ============================================================

def write_summary_sheet(wb: Workbook, books: list[dict]) -> None:
    """
    星評価ごとの集計表を別シートに書き込む。

    集計そのものは Lv02 で学んだ「辞書でグループ化」パターン。
    Excel側に数式を埋め込むこともできる（合計行で実演）。
    """
    # 新しいシートを追加
    ws = wb.create_sheet(title="評価別集計")

    # --- 星評価ごとにグループ化（Lv02 の復習） ---
    by_rating: dict[int, list[float]] = {}
    for book in books:
        by_rating.setdefault(book["rating"], []).append(book["price"])
        # setdefault: キーが無ければ初期値(空リスト)を入れてから返す。
        # if rating not in by_rating: ... の3行を1行で書ける便利メソッド

    # --- ヘッダー ---
    headers = ["星評価", "冊数", "平均価格(£)", "最高価格(£)"]
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # --- 集計行（星5 → 星1 の順に並べる） ---
    row_num = 2
    for rating in sorted(by_rating.keys(), reverse=True):
        prices = by_rating[rating]
        ws.cell(row=row_num, column=1, value=f"★{rating}")
        ws.cell(row=row_num, column=2, value=len(prices))
        ws.cell(row=row_num, column=3, value=round(sum(prices) / len(prices), 2))
        ws.cell(row=row_num, column=4, value=max(prices))
        for col_num in range(1, 5):
            ws.cell(row=row_num, column=col_num).border = THIN_BORDER
        ws.cell(row=row_num, column=3).number_format = "0.00"
        ws.cell(row=row_num, column=4).number_format = "0.00"
        row_num += 1

    # --- 合計行: Excel の数式を埋め込む ---
    # value に "=SUM(...)" のような文字列を入れると、Excelで開いたとき数式として動く
    ws.cell(row=row_num, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=row_num, column=2, value=f"=SUM(B2:B{row_num - 1})").font = Font(bold=True)
    for col_num in range(1, 5):
        ws.cell(row=row_num, column=col_num).border = THIN_BORDER

    # 列幅
    for col_num, width in enumerate([10, 8, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    print(f"  [評価別集計] {len(by_rating)} グループを集計しました")


# ============================================================
# 4. 表紙シートを作る（レポートらしさの演出）
# ============================================================

def write_cover_sheet(wb: Workbook, book_count: int) -> None:
    """レポートの概要（作成日時・件数）を先頭シートに書く。"""
    ws = wb.create_sheet(title="概要", index=0)  # index=0 で先頭に挿入

    ws["B2"] = "書籍データ レポート"
    ws["B2"].font = Font(size=16, bold=True)

    ws["B4"] = "作成日時:"
    ws["C4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["B5"] = "データ件数:"
    ws["C5"] = book_count
    ws["B6"] = "データ元:"
    ws["C6"] = "books.toscrape.com（学習用サンプル）"

    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36

    print("  [概要] 表紙シートを作成しました")


# ============================================================
# メイン処理
# ============================================================

def main() -> None:
    print("=" * 60)
    print("Lv10 - Excelレポート出力")
    print("=" * 60)

    # Workbook = Excelファイル全体を表すオブジェクト
    wb = Workbook()

    write_data_sheet(wb, BOOKS)
    write_summary_sheet(wb, BOOKS)
    write_cover_sheet(wb, len(BOOKS))

    # --- 保存 ---
    # ファイル名に日時を入れて上書きを防ぐ（Lv09 と同じパターン）
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"report_{timestamp}.xlsx"
    wb.save(output_path)

    print()
    print(f"保存完了: {output_path}")
    print("Excel で開いて、シート・書式・オートフィルタを確認してみよう。")


if __name__ == "__main__":
    main()
